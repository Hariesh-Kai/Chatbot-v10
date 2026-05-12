from __future__ import annotations

from pathlib import Path


class DocumentParser:
    """Reference parser with multi-format support.

    Supported extensions in this baseline:
    - txt, md, csv, html, xml (native)
    - pdf (via pypdf)
    - docx (via python-docx)
    - xlsx (via openpyxl)
    """

    SUPPORTED = {".txt", ".md", ".csv", ".html", ".htm", ".xml", ".pdf", ".docx", ".xlsx"}

    def parse(self, file_path: str) -> list[dict]:
        path = Path(file_path)
        ext = path.suffix.lower()
        if ext not in self.SUPPORTED:
            raise ValueError(f"Unsupported extension {path.suffix}. Supported: {', '.join(sorted(self.SUPPORTED))}")

        if ext in {".txt", ".md"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
            return self._paragraph_blocks(text)
        if ext == ".csv":
            return self._parse_csv(path)
        if ext in {".html", ".htm", ".xml"}:
            text = path.read_text(encoding="utf-8", errors="ignore")
            return self._paragraph_blocks(text)
        if ext == ".pdf":
            return self._parse_pdf(path)
        if ext == ".docx":
            return self._parse_docx(path)
        if ext == ".xlsx":
            return self._parse_xlsx(path)
        raise ValueError(f"Unhandled extension {ext}")

    @staticmethod
    def _paragraph_blocks(text: str) -> list[dict]:
        blocks: list[dict] = []
        for idx, section in enumerate([s.strip() for s in text.split("\n\n") if s.strip()], start=1):
            blocks.append({"type": "paragraph", "text": section, "page": 1, "section": f"Section {idx}"})
        return blocks

    def _parse_pdf(self, path: Path) -> list[dict]:
        try:
            from pypdf import PdfReader
        except Exception as exc:  # noqa: BLE001
            raise ValueError("PDF parsing requires dependency `pypdf`. Install project dependencies first.") from exc

        reader = PdfReader(str(path))
        blocks: list[dict] = []
        for i, page in enumerate(reader.pages, start=1):
            text = (page.extract_text() or "").strip()
            if text:
                blocks.append({"type": "paragraph", "text": text, "page": i, "section": f"Page {i}"})
        if not blocks:
            raise ValueError("PDF parsed but no text extracted (possibly scanned/image-only PDF). OCR pipeline required.")
        return blocks

    def _parse_docx(self, path: Path) -> list[dict]:
        try:
            from docx import Document
        except Exception as exc:  # noqa: BLE001
            raise ValueError("DOCX parsing requires dependency `python-docx`.") from exc

        doc = Document(str(path))
        text = "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return self._paragraph_blocks(text)

    def _parse_xlsx(self, path: Path) -> list[dict]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # noqa: BLE001
            raise ValueError("XLSX parsing requires dependency `openpyxl`.") from exc

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        blocks: list[dict] = []
        for ws in wb.worksheets:
            rows = []
            for row in ws.iter_rows(values_only=True):
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if values:
                    rows.append(" | ".join(values))
            if rows:
                blocks.append({"type": "table", "text": "\n".join(rows), "page": 1, "section": ws.title})
        return blocks

    def _parse_csv(self, path: Path) -> list[dict]:
        import csv

        rows = []
        with path.open("r", encoding="utf-8", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                vals = [c.strip() for c in row if c and c.strip()]
                if vals:
                    rows.append(" | ".join(vals))
        return [{"type": "table", "text": "\n".join(rows), "page": 1, "section": "CSV"}] if rows else []
